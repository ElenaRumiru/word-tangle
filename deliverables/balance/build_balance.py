"""
build_balance.py — Phase 5 balance table generator (Jumble prototype).

Produces the full 100-level difficulty curve as a managed system:
  - P1..P8 interpolated stepwise-monotonic between the 10 prototype anchors,
    with a sawtooth (spike every 10th level, relief right after) and a
    sublinear underlying trend.
  - difficulty_score is written as an EXCEL FORMULA in each row (weighted sum
    of normalized P1..P8), never a hardcoded number.
  - Anchor rows (1,12,23,34,45,56,67,78,89,100) carry the exact prototype
    P1..P8 and reproduce the prototype scores within +-0.15.

Outputs:
  balance-100-levels.xlsx  (Table sheet + Model sheet + Curve chart)
  balance-100-levels.csv

Run:  python build_balance.py
"""
import json
import math
import os

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
LEVELS_JSON = os.path.join(HERE, "..", "levels", "levels.json")

# --------------------------------------------------------------------------
# 1. Difficulty-score model (calibrated by minimax LP against the 10 anchors,
#    weight hierarchy constrained to be design-faithful; see balance-notes.md).
#    Features are normalized to [0,1] over each parameter's design range, so a
#    weight == that parameter's maximum contribution (in score points) at full
#    range. Score = FLOOR + sum(weight_i * norm_i).
# --------------------------------------------------------------------------
W = {
    "P1": 3.2901,  # number of words   (norm over 1..5)
    "P2": 0.9000,  # max word length   (norm over 4..8)
    "P3": 0.1000,  # rarity / zipf band(norm over 1..6)
    "P4": 2.3472,  # scramble score    (already 0..1)
    "P5": 0.3521,  # trap anagrams     (norm over 0..2)
    "P6": 0.3329,  # answer letters    (norm over 3..14)
    "P7": 0.6035,  # pun type          (encoded 0..1)
    "P8": 0.1000,  # cartoon opacity   (inverted: (5-P8)/4, 0..1)
}
FLOOR = 0.0765
PUN_ENCODE = {"direct": 0.0, "idiom": 0.33, "homophone": 0.66, "double-layer": 1.0}

# Normalization ranges (min, max) matching the feature encoding above.
NORM = {
    "P1": (1, 5),
    "P2": (4, 8),
    "P3": (1, 6),
    "P4": (0, 1),      # identity
    "P5": (0, 2),
    "P6": (3, 14),
    "P8": (1, 5),      # inverted below
}


def score_python(p):
    """Reference Python implementation of the in-cell Excel formula."""
    n1 = (p["P1"] - 1) / 4
    n2 = (p["P2"] - 4) / 4
    n3 = (p["P3"] - 1) / 5
    n4 = p["P4"]
    n5 = p["P5"] / 2
    n6 = (p["P6"] - 3) / 11
    n7 = PUN_ENCODE[p["P7"]]
    n8 = (5 - p["P8"]) / 4
    return (
        FLOOR
        + W["P1"] * n1
        + W["P2"] * n2
        + W["P3"] * n3
        + W["P4"] * n4
        + W["P5"] * n5
        + W["P6"] * n6
        + W["P7"] * n7
        + W["P8"] * n8
    )


# --------------------------------------------------------------------------
# 2. Anchors from the prototype (contract: these rows must match exactly).
# --------------------------------------------------------------------------
def load_anchors():
    data = json.load(open(LEVELS_JSON, encoding="utf-8"))
    anchors = {}
    for lvl in data["levels"]:
        d = lvl["difficulty"]
        anchors[lvl["virtualLevel"]] = {
            "P1": d["P1"], "P2": d["P2"], "P3": d["P3"], "P4": d["P4"],
            "P5": d["P5"], "P6": d["P6"], "P7": d["P7"], "P8": d["P8"],
            "score": d["score"],
        }
    return anchors


ANCHORS = load_anchors()
ANCHOR_LEVELS = sorted(ANCHORS)  # [1,12,23,34,45,56,67,78,89,100]

# Answer-word count per anchor (P6 second component) inferred from levels.json.
ANCHOR_ANSWER_WORDS = {1: 1, 12: 1, 23: 1, 34: 2, 45: 2, 56: 2, 67: 2, 78: 2, 89: 2, 100: 2}


# --------------------------------------------------------------------------
# 3. Interpolation between anchors, with sawtooth + relief.
# --------------------------------------------------------------------------
def bracket(level):
    """Return (lo_anchor, hi_anchor, t) with t in [0,1] position between them."""
    if level <= ANCHOR_LEVELS[0]:
        return ANCHOR_LEVELS[0], ANCHOR_LEVELS[0], 0.0
    if level >= ANCHOR_LEVELS[-1]:
        return ANCHOR_LEVELS[-1], ANCHOR_LEVELS[-1], 0.0
    for i in range(len(ANCHOR_LEVELS) - 1):
        lo, hi = ANCHOR_LEVELS[i], ANCHOR_LEVELS[i + 1]
        if lo <= level <= hi:
            t = (level - lo) / (hi - lo)
            return lo, hi, t
    raise RuntimeError(level)


def lerp(a, b, t):
    return a + (b - a) * t


def curve_role(level):
    """Dramaturgy tag. Every 10th = spike/milestone; level right after = relief."""
    if level in ANCHOR_LEVELS and level != 1:
        # anchor roles inherited from the prototype design
        role_map = {12: "ramp", 23: "ramp", 34: "spike", 45: "relief",
                    56: "ramp", 67: "spike", 78: "relief", 89: "ramp", 100: "milestone"}
        return role_map.get(level, "ramp")
    if level == 1:
        return "milestone"      # tutorial floor
    if level % 10 == 0:
        return "spike"
    if level % 10 == 1 and level > 1:
        return "relief"          # breather right after a spike
    return "ramp"


def sawtooth(level):
    """
    Small oscillation added on top of interpolated P4 (main continuous lever)
    so the score visibly saws: +on every 10th (spike), -right after (relief).
    Amplitude ~10-15% of local scramble range. Anchor levels get 0 (pinned).
    """
    if level in ANCHOR_LEVELS:
        return 0.0
    phase = level % 10
    if phase == 0:            # milestone spike
        return +0.07
    if phase == 1:            # relief right after
        return -0.06
    if phase in (4, 5):       # mid-block micro spike
        return +0.03
    if phase in (2, 6):       # mid-block micro relief
        return -0.03
    return 0.0


def interp_params(level):
    """Build P1..P8 for a non-anchor level by interpolating bracketing anchors."""
    if level in ANCHORS:
        a = ANCHORS[level]
        return {k: a[k] for k in ("P1", "P2", "P3", "P4", "P5", "P6", "P7", "P8")}, ANCHOR_ANSWER_WORDS[level]

    lo, hi, t = bracket(level)
    A, B = ANCHORS[lo], ANCHORS[hi]

    # Integer / stepped params: round the interpolation, stay monotone across the
    # block. Each param uses a slightly different step threshold (a per-param
    # phase offset) so that when several params change across one block their
    # discrete steps land on DIFFERENT levels instead of stacking into one big
    # jump. `stepped(a, b, t, phase)` moves the +0.5 rounding boundary by phase.
    def stepped(a, b, tt, phase):
        if a == b:
            return a
        # shift the effective position so the step fires earlier/later in block
        return int(math.floor(a + (b - a) * tt + 0.5 - phase))

    # Ascending blocks read best when steps are lightly staggered; but when the
    # block DESCENDS (a relief block, e.g. v34->v45) staggering creates wiggles.
    # Detect direction from the anchor scores and only stagger on the way up.
    ascending = B["score"] >= A["score"]
    ph = (lambda x: x) if ascending else (lambda x: 0.0)
    P1 = stepped(A["P1"], B["P1"], t, phase=ph(0.00))   # words: step at midpoint
    P2 = stepped(A["P2"], B["P2"], t, phase=ph(0.15))   # length: a touch later
    P3 = stepped(A["P3"], B["P3"], t, phase=ph(-0.15))  # rarity: a touch earlier
    P5 = stepped(A["P5"], B["P5"], t, phase=ph(0.20))   # traps: late (spike flavor)
    P6 = stepped(A["P6"], B["P6"], t, phase=ph(0.15))   # answer len: slightly late
    P8 = stepped(A["P8"], B["P8"], t, phase=ph(-0.15))  # opacity: early
    # keep integers within the two anchor values (guard against float edge)
    P1 = min(max(P1, min(A["P1"], B["P1"])), max(A["P1"], B["P1"]))
    P2 = min(max(P2, min(A["P2"], B["P2"])), max(A["P2"], B["P2"]))
    P3 = min(max(P3, min(A["P3"], B["P3"])), max(A["P3"], B["P3"]))
    P5 = min(max(P5, min(A["P5"], B["P5"])), max(A["P5"], B["P5"]))
    P6 = min(max(P6, min(A["P6"], B["P6"])), max(A["P6"], B["P6"]))
    P8 = min(max(P8, min(A["P8"], B["P8"])), max(A["P8"], B["P8"]))

    # P4 (scramble) continuous + sawtooth, clamped to [0,1].
    P4 = lerp(A["P4"], B["P4"], t) + sawtooth(level)
    P4 = round(min(1.0, max(0.0, P4)), 2)

    # P7 (pun type) is categorical: hold the LOWER anchor's pun for most of the
    # block, then adopt the higher only near its end — a discrete step, never a
    # blend. Flipping late (t>=0.7) keeps the pun change from stacking with the
    # P1/P6 steps that tend to land mid-block, smoothing the score curve.
    P7 = A["P7"] if t < 0.7 else B["P7"]

    # answer_words: step like P1's "multi-word" character.
    aw = int(round(lerp(ANCHOR_ANSWER_WORDS[lo], ANCHOR_ANSWER_WORDS[hi], t)))

    return {"P1": P1, "P2": P2, "P3": P3, "P4": P4, "P5": P5, "P6": P6, "P7": P7, "P8": P8}, aw


# --------------------------------------------------------------------------
# 4. Word-length min from max (design heuristic) & derived economy fields.
# --------------------------------------------------------------------------
def word_len_min(p2, level):
    # min length trails max by 0..2; early levels tighter (short & uniform).
    if level <= 3:
        return p2
    if p2 <= 4:
        return 4
    return max(4, p2 - 2)


def zipf_band_label(p3):
    return p3  # numeric band 1..6; label lives in notes


def free_hint_worth(level, difficulty_hint):
    """
    Soft-currency 'generosity' of the level: coins the player nets on a clean
    solve. Grows slowly with level (retention reward) but stays below hint cost
    (research: income << hint price to create scarcity). 4..10 coins.
    """
    return int(round(4 + 6 * (level / 100) ** 0.8))


def target_solve_time(level, sc):
    """
    Median expected solve time in seconds, tied to score (hypothesis).
    Calibrated so the score ladder maps onto the prototype's solve-time anchors:
    score ~0.85 -> ~25s (v1, "<30s"), score 3.4 -> ~100s (v34, "70-100"),
    score ~6.65 -> ~195s (v100, "180+").  time = 0.1 + 29.3*score.
    """
    return int(round(0.1 + 29.3 * sc))


def target_hint_usage(sc):
    """Share of players spending >=1 hint (%). Rises with score, capped ~70%."""
    return int(round(min(70, max(3, (sc - 0.8) * 12))))


# --------------------------------------------------------------------------
# 5. Milestone novelty schedule (notes column at every 10th level).
# --------------------------------------------------------------------------
MILESTONE_NOTES = {
    10: "SPIKE: first full 3-word warmup pressure; scramble sawtooth peaks",
    20: "SPIKE: first homophone pun locks in as the block standard",
    30: "SPIKE: approach to first 2-word answer; traps loom",
    40: "SPIKE: 2-word homophone answers now routine; P6 pushes to 9-10",
    50: "SPIKE: idiom puns enter rotation; longer words (P2=6)",
    60: "SPIKE: sustained trap anagrams; homophones at 12-letter answers",
    70: "SPIKE: double-layer puns become the block standard; P8 drops to 2",
    80: "SPIKE: 4-word levels harden; rarer lexicon (P3=5)",
    90: "SPIKE: near-max scramble + double-layer; exam approaches",
    100: "TOP-SPIKE (exam): 5 words, rarest lexicon, most opaque cartoon (P8=1)",
}


def notes_for(level, p, role):
    if level in MILESTONE_NOTES:
        return MILESTONE_NOTES[level]
    if level == 1:
        return "Tutorial floor: 1 word, top lexicon, no traps, direct pun"
    if level in ANCHOR_LEVELS:
        return f"ANCHOR (=prototype v{level}); pinned P1-P8"
    if role == "relief":
        return "Relief: easier scramble after the spike (breather)"
    return ""


# --------------------------------------------------------------------------
# 6. Build all 100 rows.
# --------------------------------------------------------------------------
COLUMNS = [
    "level", "words_count", "word_len_min", "word_len_max", "zipf_band",
    "scramble_score", "trap_anagrams", "answer_letters", "answer_words",
    "pun_type", "cartoon_transparency", "free_hint_worth",
    "target_solve_time_sec", "target_hint_usage_pct", "difficulty_score",
    "curve_role", "notes",
]


def build_rows():
    rows = []
    for level in range(1, 101):
        p, aw = interp_params(level)
        role = curve_role(level)
        sc = score_python(p)  # for time/economy fields (Excel recomputes score in-cell)
        row = {
            "level": level,
            "words_count": p["P1"],
            "word_len_min": word_len_min(p["P2"], level),
            "word_len_max": p["P2"],
            "zipf_band": zipf_band_label(p["P3"]),
            "scramble_score": p["P4"],
            "trap_anagrams": p["P5"],
            "answer_letters": p["P6"],
            "answer_words": aw,
            "pun_type": p["P7"],
            "cartoon_transparency": p["P8"],
            "free_hint_worth": free_hint_worth(level, sc),
            "target_solve_time_sec": target_solve_time(level, sc),
            "target_hint_usage_pct": target_hint_usage(sc),
            "difficulty_score": None,  # filled as FORMULA in xlsx
            "curve_role": role,
            "notes": notes_for(level, p, role),
            "_score_py": round(sc, 3),   # helper for CSV/verify, not a table col
        }
        rows.append(row)
    return rows


# --------------------------------------------------------------------------
# 7. Excel writer.
# --------------------------------------------------------------------------
HDR_FILL = PatternFill("solid", fgColor="2F3B52")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
ANCHOR_FILL = PatternFill("solid", fgColor="FFF2CC")
SPIKE_FONT = Font(bold=True, color="8B0000")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def col_letter(name):
    return get_column_letter(COLUMNS.index(name) + 1)


def write_xlsx(rows, path):
    wb = openpyxl.Workbook()

    # ---- Model sheet: weights + normalization, referenced by the formula ----
    ms = wb.active
    ms.title = "Model"
    ms["A1"] = "Difficulty-score model — weights on normalized P1..P8"
    ms["A1"].font = Font(bold=True, size=12)
    ms.append([])
    ms.append(["param", "weight", "norm_min", "norm_max", "note"])
    for c in range(1, 6):
        ms.cell(row=3, column=c).font = Font(bold=True)
    model_rows = [
        ("FLOOR", FLOOR, "", "", "score floor (tutorial baseline)"),
        ("P1_words", W["P1"], 1, 5, "number of words (coarse length lever)"),
        ("P2_wordlen", W["P2"], 4, 8, "max word length"),
        ("P3_zipf", W["P3"], 1, 6, "rarity band (1 common .. 6 rare)"),
        ("P4_scramble", W["P4"], 0, 1, "scramble score (already 0..1)"),
        ("P5_traps", W["P5"], 0, 2, "trap anagrams"),
        ("P6_answerlen", W["P6"], 3, 14, "answer letters"),
        ("P7_pun", W["P7"], 0, 1, "pun type encoded"),
        ("P8_opacity", W["P8"], 1, 5, "cartoon transparency, used inverted (5-P8)/4"),
    ]
    for r in model_rows:
        ms.append(r)
    # Named single cells for the formula (absolute refs).
    # FLOOR = B4 ; weights B5..B12
    ms["A15"] = "pun_type encoding"
    ms["A15"].font = Font(bold=True)
    ms.append([])
    ms.append(["direct", 0.0])
    ms.append(["idiom", 0.33])
    ms.append(["homophone", 0.66])
    ms.append(["double-layer", 1.0])
    for col in ("A", "B", "C", "D", "E"):
        ms.column_dimensions[col].width = 16
    ms["A20"] = "Score formula (per Table row): FLOOR + Σ weight_i · normalized(P_i)"
    ms["A20"].font = Font(italic=True)

    # Cell handles used inside the Table formula:
    FLOOR_REF = "Model!$B$4"
    WREF = {
        "P1": "Model!$B$5", "P2": "Model!$B$6", "P3": "Model!$B$7",
        "P4": "Model!$B$8", "P5": "Model!$B$9", "P6": "Model!$B$10",
        "P7": "Model!$B$11", "P8": "Model!$B$12",
    }

    # ---- Table sheet ----
    ws = wb.create_sheet("Table")
    ws.append(COLUMNS)
    for ci, name in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=ci)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.freeze_panes = "A2"

    cP1 = col_letter("words_count")
    cP2 = col_letter("word_len_max")
    cP3 = col_letter("zipf_band")
    cP4 = col_letter("scramble_score")
    cP5 = col_letter("trap_anagrams")
    cP6 = col_letter("answer_letters")
    cP7 = col_letter("pun_type")
    cP8 = col_letter("cartoon_transparency")

    for i, row in enumerate(rows):
        excel_row = i + 2
        vals = [row[c] for c in COLUMNS]
        ws.append(vals)
        # In-cell difficulty_score FORMULA (references Model weights + this row's P cells)
        f = (
            f"={FLOOR_REF}"
            f"+{WREF['P1']}*(({cP1}{excel_row}-1)/4)"
            f"+{WREF['P2']}*(({cP2}{excel_row}-4)/4)"
            f"+{WREF['P3']}*(({cP3}{excel_row}-1)/5)"
            f"+{WREF['P4']}*({cP4}{excel_row})"
            f"+{WREF['P5']}*(({cP5}{excel_row})/2)"
            f"+{WREF['P6']}*(({cP6}{excel_row}-3)/11)"
            f"+{WREF['P7']}*(IF({cP7}{excel_row}=\"double-layer\",1,"
            f"IF({cP7}{excel_row}=\"homophone\",0.66,"
            f"IF({cP7}{excel_row}=\"idiom\",0.33,0))))"
            f"+{WREF['P8']}*((5-{cP8}{excel_row})/4)"
        )
        ws.cell(row=excel_row, column=COLUMNS.index("difficulty_score") + 1).value = f

        # Styling: borders, anchor highlight, spike font.
        is_anchor = row["level"] in ANCHOR_LEVELS
        is_spike = row["curve_role"] in ("spike", "milestone")
        for ci in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=excel_row, column=ci)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
            if is_anchor:
                cell.fill = ANCHOR_FILL
        if is_spike:
            ws.cell(row=excel_row, column=1).font = SPIKE_FONT

    # Column widths
    widths = {
        "level": 6, "words_count": 8, "word_len_min": 8, "word_len_max": 8,
        "zipf_band": 7, "scramble_score": 9, "trap_anagrams": 8,
        "answer_letters": 9, "answer_words": 8, "pun_type": 14,
        "cartoon_transparency": 10, "free_hint_worth": 9,
        "target_solve_time_sec": 10, "target_hint_usage_pct": 10,
        "difficulty_score": 10, "curve_role": 10, "notes": 52,
    }
    for name, w in widths.items():
        ws.column_dimensions[col_letter(name)].width = w
    for name in ("difficulty_score", "scramble_score"):
        for r in range(2, 102):
            ws.cell(row=r, column=COLUMNS.index(name) + 1).number_format = "0.00"

    # Heatmap on difficulty_score (green -> yellow -> red).
    ds_col = col_letter("difficulty_score")
    ws.conditional_formatting.add(
        f"{ds_col}2:{ds_col}101",
        ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B",
        ),
    )

    # ---- Curve chart on its own sheet ----
    cs = wb.create_sheet("Curve")
    chart = LineChart()
    chart.title = "Difficulty score by level (trend + sawtooth + reliefs)"
    chart.x_axis.title = "level"
    chart.y_axis.title = "difficulty_score"
    chart.height = 11
    chart.width = 26
    data = Reference(ws, min_col=COLUMNS.index("difficulty_score") + 1,
                     min_row=1, max_row=101)
    cats = Reference(ws, min_col=1, min_row=2, max_row=101)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.y_axis.majorGridlines = None
    s = chart.series[0]
    s.smooth = False
    cs.add_chart(chart, "B2")
    cs["B22"] = "Spikes at every 10th level; reliefs at the level right after; sublinear trend underneath."

    wb.save(path)


def write_csv(rows, path):
    import csv
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(COLUMNS)
        for row in rows:
            out = []
            for c in COLUMNS:
                if c == "difficulty_score":
                    out.append(row["_score_py"])  # CSV carries the computed value
                else:
                    out.append(row[c])
            w.writerow(out)


def main():
    rows = build_rows()
    xlsx = os.path.join(HERE, "balance-100-levels.xlsx")
    csvp = os.path.join(HERE, "balance-100-levels.csv")
    write_xlsx(rows, xlsx)
    write_csv(rows, csvp)

    # Sanity: anchor rows reproduce prototype scores within +-0.15
    print("Anchor check (python recompute):")
    worst = 0.0
    for lvl in ANCHOR_LEVELS:
        row = rows[lvl - 1]
        got = row["_score_py"]
        want = ANCHORS[lvl]["score"]
        err = got - want
        worst = max(worst, abs(err))
        flag = "OK" if abs(err) <= 0.15 else "FAIL"
        print(f"  v{lvl:>3}: want={want:.2f} got={got:.3f} err={err:+.3f} [{flag}]")
    print(f"worst anchor err = {worst:.4f}")
    print(f"wrote {xlsx}")
    print(f"wrote {csvp}")


if __name__ == "__main__":
    main()
