"""
verify_balance.py — independent verification of balance-100-levels.xlsx.

Checks:
  1. Table sheet has exactly 100 data rows.
  2. Every difficulty_score cell is a FORMULA (string starting with '=').
  3. Re-evaluates the score model in Python from each row's P1..P8 cells and
     confirms the 10 anchor rows (1,12,23,34,45,56,67,78,89,100) reproduce the
     prototype scores within +-0.15.
  4. Curve shape: score is broadly increasing (early < late) and shows a
     sawtooth (a dip exists right after at least one milestone spike).

Run:  python verify_balance.py
Exit code 0 = all checks pass.
"""
import json
import os
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "balance-100-levels.xlsx")
LEVELS_JSON = os.path.join(HERE, "..", "levels", "levels.json")

W = {"P1": 3.2901, "P2": 0.9000, "P3": 0.1000, "P4": 2.3472,
     "P5": 0.3521, "P6": 0.3329, "P7": 0.6035, "P8": 0.1000}
FLOOR = 0.0765
PUN = {"direct": 0.0, "idiom": 0.33, "homophone": 0.66, "double-layer": 1.0}
ANCHOR_LEVELS = [1, 12, 23, 34, 45, 56, 67, 78, 89, 100]
TOL = 0.15


def score(p):
    return (
        FLOOR
        + W["P1"] * (p["P1"] - 1) / 4
        + W["P2"] * (p["P2"] - 4) / 4
        + W["P3"] * (p["P3"] - 1) / 5
        + W["P4"] * p["P4"]
        + W["P5"] * p["P5"] / 2
        + W["P6"] * (p["P6"] - 3) / 11
        + W["P7"] * PUN[p["P7"]]
        + W["P8"] * (5 - p["P8"]) / 4
    )


def main():
    fails = []

    # data_only=False -> we see FORMULAS, not cached values.
    wb = openpyxl.load_workbook(XLSX, data_only=False)
    assert "Table" in wb.sheetnames, "Table sheet missing"
    ws = wb["Table"]
    header = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}

    # 1. row count
    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] is not None]
    n = len(data_rows)
    print(f"[1] data rows: {n}  -> {'OK' if n == 100 else 'FAIL'}")
    if n != 100:
        fails.append("row count != 100")

    # 2. difficulty_score is a formula in every row
    ds_col = idx["difficulty_score"] + 1
    formula_ok = 0
    for r in range(2, 2 + n):
        v = ws.cell(row=r, column=ds_col).value
        if isinstance(v, str) and v.startswith("="):
            formula_ok += 1
    print(f"[2] formula cells in difficulty_score: {formula_ok}/{n} "
          f"-> {'OK' if formula_ok == n else 'FAIL'}")
    if formula_ok != n:
        fails.append("not all difficulty_score cells are formulas")

    # 3. recompute from P-cells, check anchors
    def rowp(r):
        g = lambda name: ws.cell(row=r, column=idx[name] + 1).value
        return {
            "P1": g("words_count"), "P2": g("word_len_max"), "P3": g("zipf_band"),
            "P4": g("scramble_score"), "P5": g("trap_anagrams"),
            "P6": g("answer_letters"), "P7": g("pun_type"),
            "P8": g("cartoon_transparency"),
        }

    anchors = {L["virtualLevel"]: L["difficulty"]["score"]
               for L in json.load(open(LEVELS_JSON, encoding="utf-8"))["levels"]}
    proto_p = {L["virtualLevel"]: L["difficulty"]
               for L in json.load(open(LEVELS_JSON, encoding="utf-8"))["levels"]}

    print("[3] anchor reproduction (recomputed formula vs prototype score):")
    worst = 0.0
    for lvl in ANCHOR_LEVELS:
        p = rowp(lvl + 1)  # excel row = level+1
        s = score(p)
        want = anchors[lvl]
        err = s - want
        worst = max(worst, abs(err))
        ok = abs(err) <= TOL
        # also confirm P1..P8 in the table match the prototype exactly
        pp = proto_p[lvl]
        pmatch = all(abs((p[k] if not isinstance(p[k], str) else 0) -
                         (pp[k] if not isinstance(pp[k], str) else 0)) < 1e-9
                     for k in ("P1", "P2", "P3", "P4", "P5", "P6", "P8")) \
            and p["P7"] == pp["P7"]
        tag = "OK" if ok and pmatch else "FAIL"
        print(f"    v{lvl:>3}: want={want:.2f} got={s:.3f} err={err:+.3f} "
              f"P-match={pmatch} [{tag}]")
        if not ok:
            fails.append(f"anchor v{lvl} score out of tolerance")
        if not pmatch:
            fails.append(f"anchor v{lvl} P1-P8 mismatch vs prototype")
    print(f"    worst anchor err = {worst:.4f} (tol {TOL})")

    # 4. curve shape: increasing overall + sawtooth (dip after a spike)
    scores = [score(rowp(r)) for r in range(2, 2 + n)]
    early = sum(scores[:10]) / 10
    late = sum(scores[-10:]) / 10
    inc_ok = late > early + 2.0
    print(f"[4a] trend increasing: early avg {early:.2f} < late avg {late:.2f} "
          f"-> {'OK' if inc_ok else 'FAIL'}")
    if not inc_ok:
        fails.append("curve not increasing")

    dips = 0
    for L in range(10, 100, 10):
        # spike at level L (index L-1), relief at L+1 (index L)
        if L < 100 and scores[L] < scores[L - 1]:
            dips += 1
    saw_ok = dips >= 3
    print(f"[4b] sawtooth reliefs after spikes: {dips} milestones dip -> "
          f"{'OK' if saw_ok else 'FAIL'}")
    if not saw_ok:
        fails.append("no clear sawtooth")

    print()
    if fails:
        print("RESULT: FAIL")
        for f in fails:
            print("  -", f)
        sys.exit(1)
    print("RESULT: ALL CHECKS PASS")
    sys.exit(0)


if __name__ == "__main__":
    main()
